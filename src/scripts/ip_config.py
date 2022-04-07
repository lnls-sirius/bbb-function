
from openpyxl import load_workbook
import subprocess

# Checks  if flash memory  is mounted 
def pendrive():
	p = subprocess.run(['lsblk'],stdout= subprocess.PIPE)
	stdout = p.stdout
	return stdout.decode('ISO-8859-1') if p.returncode == 0 else ''

#Mount pendrive
def mount_pendrive():
	p  = subprocess.run(['mount','/dev/sda1','/mnt/USB'],stdout = subprocess.PIPE)
	stdout = p.stdout
	return stdout.decode('ISO-8859-1') if p.returncode == 0 else ''


#Checks flash memory  is connected
def check_devices():
	p = subprocess.run(['ls','/dev'],stdout = subprocess.PIPE)
	stdout = p.stdout
	return  stdout.decode('ISO-8859-1') if p.returncode == 0 else ''

# Checks what kind of file this in the flash memory
def read_pendrive():
	p = subprocess.run(['ls','/mnt/USB'],stdout= subprocess.PIPE)
	stdout = p.stdout
	return stdout.decode('ISO-8859-1') if p.returncode == 0 else ''

#Finds name service and changes ip 
def get_name_service():
    p = subprocess.run(['connmanctl', 'services'], stdout=subprocess.PIPE)
    stdout  = p.stdout
    return stdout.decode('ISO-8859-1') if p.returncode == 0 else ''

def change_ip(service,new_ip,mask,gateway):
	p = subprocess.run(['connmanctl','config',service,'--ipv4','manual',new_ip,mask,gateway],stdout = subprocess.PIPE)

	name_service = get_name_service()
	service_n =  name_service.split(" ")
	service = service_n[17][:-1]
	change_ip(service,new_ip,mask,gateway)



# Starts checking  for a flash memory
find_devices =  check_devices()

if(find_devices.find('sda') == -1 and find_devices.find('sdb') == -1):
	print("Pendrive  not detected")
	exit()
else:

	devices = pendrive()


	result = devices.find('/mnt/USB')

	if(result == -1):
		mount_pendrive()

	pendrive_files  = read_pendrive()

	if(pendrive_files.find('AutoConfig.xlsx') == -1 and pendrive_files.find('AutoConfig.txt') == -1):
		print("No files detected")

	elif(pendrive_files.find('AutoConfig.xlsx') ==-1  and pendrive_files.find('AutoConfig.txt') !=  -1):
		file = open('/mnt/USB/AutoConfig.txt')
		lines = file.readlines()
		bbb = lines[0].split("\n")
		print(bbb[0])
		bbb_name = bbb[0]


		

	else:
		wb = load_workbook('mnt/USB/AutoConfig.xlsx')
		ws = wb.active
		new_ip = (ws['A2']).value
		mask = (ws['B2']).value	
		gateway = (ws['C2']).value

		print("Changing IP to:")
		print(new_ip,mask,gateway)




